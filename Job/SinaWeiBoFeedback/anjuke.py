# -*- coding: utf-8 -*-
# @Time    : 2017/9/19 21:36
# @Author  : 蛇崽
# @Email   : 17193337679@163.com
# @File    : anjuke.py 安居客房产网
import requests
import re
from bs4 import BeautifulSoup
import csv
import time
import threading
from lxml import etree
from selenium import webdriver
from openpyxl import Workbook

num0 = 1  # 用来计数，计算爬取的书一共有多少本
baseurl = 'https://www.anjuke.com/sy-city.html'

wb = Workbook()
ws = wb.active
ws.title = '安居客'
ws.cell(row=1, column=1).value = '城市名称'
ws.cell(row=1, column=2).value = '城市链接'

def gethtml():
    chromedriver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
    browser = webdriver.Chrome(chromedriver)
    browser.get(baseurl)
    time.sleep(5)
    js = 'window.scrollBy(0,3000)'
    browser.execute_script(js)
    js = 'window.scrollBy(0,5000)'
    browser.execute_script(js)
    html = browser.page_source
    return html


def saveinfos(authorother):
    global num0
    nums = 0
    for ver_info in authorother:
        num0 = num0 + 1
        ws.cell(row=num0, column=1).value = ver_info[0]
        ws.cell(row=num0, column=2).value = ver_info[1]
        nums += 1
        print('爬取成功 ' + str(nums))
    wb.save('安居客' + '.xlsx')
    pass


def parseHotBook(html):
    # 作者 （豆瓣用户，简书）
    print(html)
    print('*'*20)
    # commentlist = html.xpath("/html/body/div[3]/div")
    # 作者 （豆瓣用户，简书）
    regAuthor = r'.*?<a href="(.*?)</a>'
    reg_author = re.compile(regAuthor)
    authorother = re.findall(reg_author, html)

    global num0
    nums = 0

    for info in authorother:
        verinfo = info.split('">')
        print(verinfo[0],verinfo[1].replace('class="hot',''))

        num0 = num0 + 1
        name = verinfo[0]
        link = verinfo[1].replace('class="hot','')
        ws.cell(row=num0, column=1).value = name
        ws.cell(row=num0, column=2).value = link
    wb.save('安居客2' + '.xlsx')
    print('爬取成功')
html = gethtml()
parseHotBook(html)
















